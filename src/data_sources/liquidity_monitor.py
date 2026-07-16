"""AI-bubble liquidity & momentum monitor.

Tracks the five indicators from the "AI bubble mechanics" thesis and scores each
by its trailing-window percentile (auto-recalibrating) while surfacing the
creator's absolute thresholds as reference context.

Data sources (all free / already-provisioned):
  - FRED           -> reserves (WRESBAL), SOFR, IORB, TGA (WTREGEN), ON RRP
  - NY Fed markets -> Standing Repo Facility take-up
  - SEC EDGAR      -> quarterly cloud-provider capex (companyconcept XBRL)
  - yfinance       -> options skew (best-effort; degrades to last stored value)

Design notes
------------
* Every fetcher returns a list of ``{"date": "YYYY-MM-DD", "value": float}`` points
  (oldest -> newest) plus metadata, so the same evaluator handles all metrics.
* ``LiquidityMonitor.snapshot()`` is the single entry point the API/scheduler call.
* Nothing here raises on a single-source failure; a failed metric is returned with
  ``status="unavailable"`` so one dead API never takes down the whole dashboard.
"""

from __future__ import annotations

import json
import time
import urllib.request
from datetime import date, datetime
from typing import Any

import yaml

from src.config import Keys, Paths
from src.utils.cache import DataCache
from src.utils.logger import setup_logger

logger = setup_logger("liquidity_monitor")

CONFIG_PATH = Paths.ROOT / "configs" / "bubble_monitor.yaml"
SEC_UA = Keys.SEC_USER_AGENT or "FE-Analyst research adamxyz96@gmail.com"


def _load_config() -> dict:
    with open(CONFIG_PATH) as f:
        return yaml.safe_load(f)


def _http_json(url: str, ua: str = "Mozilla/5.0", timeout: int = 30) -> Any:
    req = urllib.request.Request(url, headers={"User-Agent": ua})
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return json.load(resp)


def _http_csv(
    url: str, ua: str = "Mozilla/5.0", timeout: int = 30, retries: int = 3
) -> list[str]:
    """Fetch a CSV with retries — CDN endpoints (e.g. CBOE) occasionally reset."""
    last_err: Exception | None = None
    for attempt in range(retries):
        try:
            req = urllib.request.Request(url, headers={"User-Agent": ua})
            with urllib.request.urlopen(req, timeout=timeout) as resp:
                return resp.read().decode("utf-8", "ignore").splitlines()
        except Exception as e:  # noqa: BLE001
            last_err = e
            if attempt < retries - 1:
                time.sleep(1.5 * (attempt + 1))
    raise last_err  # type: ignore[misc]


def _percentile_rank(values: list[float], current: float) -> float | None:
    """Return the percentile (0-100) of ``current`` within ``values``."""
    clean = [v for v in values if v is not None]
    if len(clean) < 5:
        return None
    below = sum(1 for v in clean if v <= current)
    return round(100.0 * below / len(clean), 1)


class LiquidityMonitor:
    """Fetch, score, and package the AI-bubble indicator set."""

    def __init__(self, config: dict | None = None):
        self.cfg = config or _load_config()
        self.window = int(self.cfg.get("percentile_window_days", 504))
        self.cache = DataCache("liquidity_monitor")
        # Lazy MacroDataClient (wraps FRED + caching).
        from src.data_sources.macro_data import MacroDataClient

        self._macro = MacroDataClient()

    # ------------------------------------------------------------------ FRED
    def _fred_series(self, series_id: str, start: str = "2015-01-01") -> list[dict]:
        s = self._macro.get_fred_series(series_id, start=start)
        if s is None or getattr(s, "empty", True):
            return []
        s = s.dropna()
        return [
            {"date": idx.strftime("%Y-%m-%d"), "value": float(val)}
            for idx, val in s.items()
        ]

    def _fetch_fred(self, m: dict) -> list[dict]:
        return self._fred_series(m["series"])

    def _fetch_fred_spread(self, m: dict) -> list[dict]:
        """SOFR - IORB, in basis points, aligned on common dates."""
        a = {p["date"]: p["value"] for p in self._fred_series(m["series_a"])}
        b = {p["date"]: p["value"] for p in self._fred_series(m["series_b"])}
        common = sorted(set(a) & set(b))
        return [{"date": d, "value": round((a[d] - b[d]) * 100, 2)} for d in common]

    # --------------------------------------------------------------- NY Fed SRF
    def _fetch_nyfed_srf(self, m: dict) -> list[dict]:
        """Daily total accepted amount across Standing Repo Facility operations.

        The NY Fed repo endpoint returns individual operations; SRF ops are
        tagged in the operation metadata. We sum accepted amounts per day so a
        spike in take-up is visible even across multiple same-day operations.
        """
        n = int(m.get("lookback_ops", 40))
        url = f"https://markets.newyorkfed.org/api/rp/repo/all/results/last/{n}.json"
        try:
            data = _http_json(url)
        except Exception as e:  # noqa: BLE001
            logger.warning("NY Fed SRF fetch failed: %s", e)
            return []
        ops = data.get("repo", {}).get("operations", [])
        per_day: dict[str, float] = {}
        for op in ops:
            d = op.get("operationDate")
            if not d:
                continue
            amt = op.get("totalAmtAccepted") or 0
            per_day[d] = per_day.get(d, 0.0) + float(amt)
        return [{"date": d, "value": per_day[d]} for d in sorted(per_day)]

    # ---------------------------------------------------------------- SEC capex
    def _sec_quarterly_capex(self, cik: str, concepts: list[str]) -> dict[str, float]:
        """Return {quarter-end -> single-quarter capex USD} for one company.

        Handles both reporting styles:
          * standalone-quarter reporters (e.g. MSFT) — ~90-day periods used directly
          * year-to-date reporters (e.g. AMZN) — quarterly value is derived by
            differencing consecutive cumulative periods within the same fiscal year.

        Companies migrate XBRL tags over time (AMZN switched to
        ``PaymentsToAcquireProductiveAssets``), so we pick the concept whose data
        extends furthest — i.e. the one actually in use today.
        """
        best_units: list[dict] = []
        best_latest = ""
        for concept in concepts:
            url = (
                f"https://data.sec.gov/api/xbrl/companyconcept/"
                f"CIK{cik}/us-gaap/{concept}.json"
            )
            try:
                data = _http_json(url, ua=SEC_UA)
            except Exception:  # noqa: BLE001
                continue
            units = data.get("units", {}).get("USD", [])
            if not units:
                continue
            latest = max(u["end"] for u in units)
            if latest > best_latest:
                best_latest, best_units = latest, units
        return self._derive_quarters(best_units)

    @staticmethod
    def _derive_quarters(units: list[dict]) -> tuple[dict[str, float], dict[str, str]]:
        """Reduce raw XBRL USD facts to single-quarter capex keyed by period end.

        Returns (quarters {end->value}, filed {end->SEC filing date})."""

        def days(u: dict) -> int:
            return (date.fromisoformat(u["end"]) - date.fromisoformat(u["start"])).days

        # Dedupe overlapping (start,end) periods, keeping the latest-filed value.
        by_period: dict[tuple[str, str], dict] = {}
        for u in units:
            if u.get("form") not in ("10-Q", "10-K"):
                continue
            key = (u["start"], u["end"])
            prev = by_period.get(key)
            if prev is None or u.get("filed", "") > prev.get("filed", ""):
                by_period[key] = u
        entries = list(by_period.values())

        quarters: dict[str, float] = {}
        filed: dict[str, str] = {}
        # 1) Standalone ~quarter periods are used directly.
        for u in entries:
            if 80 <= days(u) <= 100:
                quarters[u["end"]] = float(u["val"])
                filed[u["end"]] = u.get("filed", "")
        # 2) YTD reporters: within a shared fiscal-year start, difference cumulatives.
        groups: dict[str, list[dict]] = {}
        for u in entries:
            groups.setdefault(u["start"], []).append(u)
        for group in groups.values():
            group.sort(key=lambda u: u["end"])
            prev_val: float | None = None
            for u in group:
                if days(u) <= 100:
                    prev_val = float(u["val"])  # anchor = the standalone Q1
                    continue
                if prev_val is not None:
                    q = float(u["val"]) - prev_val
                    if u["end"] not in quarters and q > 0:
                        quarters[u["end"]] = q
                        filed[u["end"]] = u.get("filed", "")
                prev_val = float(u["val"])
        return quarters, filed

    def _fetch_sec_capex(self, m: dict) -> list[dict]:
        """Cloud-provider capex -> sequential (QoQ) growth-rate series.

        The stored *value* is the average QoQ growth rate (%) across the
        hyperscalers that reported. We compute each company's own quarter-over-
        quarter growth first, then average — this tolerates the companies'
        differing fiscal calendars far better than requiring a shared frame.

        Also exposed: ``capex_usd`` = total capex of reporting companies that
        quarter (absolute level, for context), and ``companies`` = coverage count.
        The evaluator flags the 2nd derivative (a rolling *deceleration* of this
        rate) — the creator's "ultimate warning".
        """
        concepts = m["concepts"]
        min_cos = 2  # need at least this many companies for a meaningful average

        # Per-company: {quarter_end -> capex}, and per-company QoQ growth.
        per_qtr_growth: dict[str, list[float]] = {}
        per_qtr_capex: dict[str, float] = {}
        per_qtr_count: dict[str, int] = {}
        per_qtr_filed: dict[str, str] = {}  # latest SEC filing date per quarter
        for name, meta in m["companies"].items():
            q, filed = self._sec_quarterly_capex(meta["cik"], concepts)
            if not q:
                logger.warning("No capex data for %s", name)
                continue
            dates = sorted(q)
            for i, d in enumerate(dates):
                per_qtr_capex[d] = per_qtr_capex.get(d, 0.0) + q[d]
                per_qtr_count[d] = per_qtr_count.get(d, 0) + 1
                per_qtr_filed[d] = max(per_qtr_filed.get(d, ""), filed.get(d, ""))
                if i == 0:
                    continue
                prev = q[dates[i - 1]]
                if prev > 0:
                    growth = 100.0 * (q[d] - prev) / prev
                    per_qtr_growth.setdefault(d, []).append(growth)

        series = []
        for d in sorted(per_qtr_growth):
            growths = per_qtr_growth[d]
            if len(growths) < min_cos:
                continue
            series.append(
                {
                    "date": d,
                    "value": round(sum(growths) / len(growths), 2),
                    "capex_usd": round(per_qtr_capex.get(d, 0.0), 0),
                    "companies": per_qtr_count.get(d, len(growths)),
                    "filed": per_qtr_filed.get(d, ""),
                }
            )
        return series

    # ------------------------------------------------------------- options skew
    def _fetch_cboe_skew(self, m: dict) -> list[dict]:
        """CBOE SKEW Index — the market-standard options-skew gauge.

        Sourced from CBOE's own published daily-price CSV (no API key, no rate
        limits). Measures the priced cost of OTM puts relative to ATM: a high
        reading = heavy tail/crash-hedging demand; a low/falling reading = cheap
        downside protection = the complacency that accompanies melt-ups (the
        thesis's "calls richer than puts"). Reliable replacement for scraped
        option-chain risk-reversals.
        """
        url = m.get(
            "url",
            "https://cdn.cboe.com/api/global/us_indices/daily_prices/SKEW_History.csv",
        )
        try:
            lines = _http_csv(url)
        except Exception as e:  # noqa: BLE001
            logger.warning("CBOE SKEW fetch failed: %s", e)
            return []
        out = []
        for row in lines[1:]:  # skip header "DATE,SKEW"
            parts = row.split(",")
            if len(parts) < 2 or not parts[1].strip():
                continue
            try:
                d = datetime.strptime(parts[0].strip(), "%m/%d/%Y").strftime("%Y-%m-%d")
                out.append({"date": d, "value": round(float(parts[1]), 2)})
            except ValueError:
                continue
        return out

    _FETCHERS = {
        "fred": "_fetch_fred",
        "fred_spread": "_fetch_fred_spread",
        "nyfed_srf": "_fetch_nyfed_srf",
        "sec_capex": "_fetch_sec_capex",
        "cboe_skew": "_fetch_cboe_skew",
    }

    # -------------------------------------------------------------- evaluation
    def _evaluate(self, key: str, m: dict, series: list[dict]) -> dict:
        """Score one metric: current value, percentile, status, references."""
        result: dict[str, Any] = {
            "key": key,
            "label": m["label"],
            "unit": m.get("unit"),
            "direction": m["direction"],
            "note": m.get("note", ""),
            "reference_yellow": m.get("reference_yellow"),
            "reference_red": m.get("reference_red"),
            "context_only": bool(m.get("context_only", False)),
            "informational": bool(m.get("informational", False)),
            "series": series[-self.window :] if series else [],
        }
        if not series:
            result.update(status="unavailable", current=None, percentile=None)
            return result

        latest = series[-1]
        current = latest["value"]
        window_vals = [p["value"] for p in series[-self.window :]]
        pct = _percentile_rank(window_vals, current)

        result.update(
            current=current,
            as_of=latest["date"],
            percentile=pct,
            capex_usd=latest.get("capex_usd"),
        )
        result["status"] = self._status(m, current, pct)
        return result

    def _status(self, m: dict, current: float, pct: float | None) -> str:
        """Combine the creator's absolute thresholds with percentile extremes."""
        direction = m["direction"]
        y, r = m.get("reference_yellow"), m.get("reference_red")
        alerts = self.cfg.get("alerts", {})
        y_pct = alerts.get("yellow_percentile", 90)
        r_pct = alerts.get("red_percentile", 97)

        status = "normal"
        if direction == "up_bad":
            if r is not None and current >= r:
                status = "red"
            elif y is not None and current >= y:
                status = "yellow"
            if pct is not None:
                if pct >= r_pct:
                    status = "red"
                elif pct >= y_pct and status == "normal":
                    status = "yellow"
        else:  # down_bad — low values are the danger
            if r is not None and current <= r:
                status = "red"
            elif y is not None and current <= y:
                status = "yellow"
            if pct is not None:
                if pct <= (100 - r_pct):
                    status = "red"
                elif pct <= (100 - y_pct) and status == "normal":
                    status = "yellow"
        return status

    @staticmethod
    def _srf_is_stress(srf, reserves, sofr, c) -> bool:
        """An SRF spike only signals a structural top when CORROBORATED by reserve
        or funding stress. A lone repo blip while reserves are ample is NOT a top —
        this matches the creator's 'resonance' caveat and prevents false alarms from
        large-but-routine Full-Allotment SRF operations (which our raw sum can read
        as a $100B 'spike')."""
        if srf is None or srf < c.get("srf_spike", 5000000):
            return False
        reserves_draining = reserves is not None and reserves < c.get("reserves_yellow_break", 2900000)
        sofr_positive = sofr is not None and sofr > c.get("sofr_positive_bps", 0.0)
        return reserves_draining or sofr_positive

    @staticmethod
    def _capex_rolling_over(series: list[dict]) -> bool:
        """Seasonally-aware capex deceleration: latest QoQ growth below the same
        quarter a year ago (~4 quarters back). Falls back to a simple 2nd
        derivative if there isn't a full year of history."""
        vals = [p["value"] for p in series if p.get("value") is not None]
        if len(vals) >= 5:
            return vals[-1] < vals[-5]
        if len(vals) >= 2:
            return vals[-1] < vals[-2]
        return False

    # ---------------------------------------------------------- confluence
    def _evaluate_confluence(self, metrics: dict) -> dict:
        """Evaluate the creator's AND/OR combo alerts across metrics.

        YELLOW = SOFR positive AND reserves<$2.9T AND TGA→$1T
        RED    = (SOFR>3bps AND reserves<$2.8T)  OR  SRF spike
        TOP    = (yellow or red liquidity) AND cloud-capex QoQ rolling over
        """
        c = self.cfg.get("confluence", {})

        def cur(key):
            return metrics.get(key, {}).get("current")

        sofr = cur("sofr_iorb_spread")
        reserves = cur("reserves")
        tga = cur("tga")
        srf = cur("srf_usage")

        # capex rolling over — seasonally aware (capex is highly seasonal, so a raw
        # Q/Q dip isn't a rollover; compare latest QoQ to the same quarter a year ago).
        capex_series = metrics.get("cloud_capex_accel", {}).get("series", [])
        capex_decel = self._capex_rolling_over(capex_series)

        # SOFR-IORB must be positive on CONSECUTIVE days (creator: "连续转正"),
        # not a single-day blip — a lone positive print is not a yellow signal.
        sofr_series = metrics.get("sofr_iorb_spread", {}).get("series", [])
        sofr_days = int(c.get("sofr_positive_days", 2))
        sofr_thr = c.get("sofr_positive_bps", 0.0)
        sofr_sustained = (
            len(sofr_series) >= sofr_days
            and all(p["value"] > sofr_thr for p in sofr_series[-sofr_days:])
        )

        def cond(label, met, detail):
            return {"label": label, "met": bool(met), "detail": detail}

        yellow = [
            cond(f"SOFR-IORB positive {sofr_days}d", sofr_sustained,
                 f"{sofr:+.1f}bp" if sofr is not None else "n/a"),
            cond("Reserves breaking $2.9T", reserves is not None and reserves < c.get("reserves_yellow_break", 2900000),
                 f"${reserves/1e6:.2f}T" if reserves is not None else "n/a"),
            cond("TGA → $1T", tga is not None and tga >= c.get("tga_approaching", 950000),
                 f"${tga/1e6:.2f}T" if tga is not None else "n/a"),
        ]
        red = [
            cond("SOFR>3bps AND reserves<$2.8T",
                 sofr is not None and reserves is not None
                 and sofr > c.get("sofr_red_bps", 3.0) and reserves < c.get("reserves_red_break", 2800000),
                 f"{sofr:+.1f}bp / ${reserves/1e6:.2f}T" if sofr is not None and reserves is not None else "n/a"),
            cond("SRF spike + reserves draining", self._srf_is_stress(srf, reserves, sofr, c),
                 (f"${srf/1e6:.2f}B"
                  + (" (reserves ample — not corroborated)"
                     if srf is not None and srf >= c.get("srf_spike", 5000000)
                     and not self._srf_is_stress(srf, reserves, sofr, c) else ""))
                 if srf is not None else "n/a"),
            cond("Reserves < $2.5T (Rule 1: liquidate)",
                 reserves is not None and reserves < c.get("reserves_liquidate", 2500000),
                 f"${reserves/1e6:.2f}T" if reserves is not None else "n/a"),
        ]
        top = [
            cond("Cloud-capex QoQ rolling over", capex_decel,
                 f"{capex_series[-1]['value']:+.1f}% (prev {capex_series[-2]['value']:+.1f}%)"
                 if len(capex_series) >= 2 else "n/a"),
        ]

        yellow_met = all(x["met"] for x in yellow)
        red_met = any(x["met"] for x in red)
        top_met = (yellow_met or red_met) and capex_decel

        if top_met:
            level, headline = "top", "Structural top: liquidity stress AND capex rolling over (resonance)"
        elif red_met:
            level, headline = "red", "Red: funding-stress combo met — probable structural top"
        elif yellow_met:
            level, headline = "yellow", "Yellow: liquidity-tightening combo forming"
        else:
            level, headline = "normal", "Normal: no confluence combo met (not even yellow)"

        return {
            "level": level,
            "headline": headline,
            "yellow": yellow,
            "red": red,
            "top": top,
            "capex_rolling_over": capex_decel,
        }

    # ----------------------------------------------------------- top model
    def _fetch_fred_ratio(self, m: dict) -> tuple[float | None, str | None]:
        num = self._macro.get_fred_series(m["series_num"])
        den = self._macro.get_fred_series(m["series_den"])
        if num is None or den is None or num.empty or den.empty:
            return None, None
        scale = float(m.get("scale", 1.0))
        val = round(100.0 * (float(num.iloc[-1]) * scale) / float(den.iloc[-1]), 1)
        as_of = min(num.index[-1], den.index[-1]).strftime("%Y-%m-%d")  # oldest of the two
        return val, as_of

    def _fetch_fred_yoy_diff(self, m: dict) -> tuple[float | None, str | None]:
        """YoY% of series_a minus YoY% of series_b (date-based lookback)."""
        import pandas as pd

        as_ofs = []

        def yoy(sid):
            s = self._macro.get_fred_series(sid)
            if s is None or s.empty:
                return None
            s = s.dropna().sort_index()
            as_ofs.append(s.index[-1])
            cur = float(s.iloc[-1])
            past = s.asof(s.index[-1] - pd.DateOffset(years=1))  # value ~12mo ago
            if past is None or pd.isna(past) or float(past) == 0:
                return None
            return 100.0 * (cur - float(past)) / float(past)

        a, b = yoy(m["series_a"]), yoy(m["series_b"])
        if a is None or b is None:
            return None, None
        return round(a - b, 1), (min(as_ofs).strftime("%Y-%m-%d") if as_ofs else None)

    def _fetch_ssga_top10(self, m: dict) -> tuple[float | None, str | None]:
        """Sum of the top-10 holdings' weights from State Street's SPY file (official)."""
        import io

        try:
            import openpyxl
        except ImportError:
            return None, None
        url = m.get("url")
        try:
            raw = urllib.request.urlopen(
                urllib.request.Request(url, headers={"User-Agent": "Mozilla/5.0"}), timeout=30
            ).read()
            wb = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
            ws = wb.active
            rows = [[c.value for c in r] for r in ws.iter_rows(max_row=12)]
        except Exception as e:  # noqa: BLE001
            logger.warning("SSGA top-10 fetch failed: %s", e)
            return None, None
        # As-of date lives in a "As of DD-Mon-YYYY" cell near the top.
        as_of = None
        for r in rows:
            for cell in r:
                if isinstance(cell, str) and "As of" in cell:
                    try:
                        as_of = datetime.strptime(
                            cell.split("As of")[-1].strip(), "%d-%b-%Y"
                        ).strftime("%Y-%m-%d")
                    except ValueError:
                        pass
        hdr_i = next((i for i, r in enumerate(rows)
                      if r and any(str(c).strip() == "Weight" for c in r if c)), None)
        if hdr_i is None:
            return None, as_of
        wcol = [j for j, c in enumerate(rows[hdr_i]) if str(c).strip() == "Weight"][0]
        weights = []
        for r in ws.iter_rows(min_row=hdr_i + 2, values_only=True):
            try:
                w = float(str(r[wcol]).replace("%", ""))
            except (TypeError, ValueError):
                continue
            if w > 0:
                weights.append(w)
            if len(weights) >= 10:
                break
        if len(weights) < 10:
            return None, as_of
        return round(sum(weights[:10]), 2), as_of

    _MONTHS = ["", "Jan", "Feb", "Mar", "Apr", "May", "Jun",
               "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

    @staticmethod
    def _quarter_label(end_str: str) -> str:
        try:
            d = date.fromisoformat(end_str)
        except (ValueError, TypeError):
            return end_str
        return f"Q{(d.month - 1) // 3 + 1} {d.year}"

    def _next_report_window(self) -> str:
        """Human label for the next hyperscaler earnings window (they file ~day 25)."""
        months = sorted(self.cfg.get("earnings_calendar", {}).get("report_months", [1, 4, 7, 10]))
        now = datetime.utcnow()
        for mo in months:
            if mo > now.month or (mo == now.month and now.day < 25):
                return f"late {self._MONTHS[mo]} {now.year}"
        return f"late {self._MONTHS[months[0]]} {now.year + 1}"

    def _liquidity_calendar(self) -> dict | None:
        """Active/imminent Treasury liquidity-drain window (context, not a sell)."""
        cfg = self.cfg.get("liquidity_calendar", {})
        lead = int(cfg.get("lead_days", 30))
        today = datetime.utcnow().date()
        for e in cfg.get("events", []):
            try:
                start = date.fromisoformat(e["start"])
                peak = date.fromisoformat(e["peak"])
            except (ValueError, KeyError):
                continue
            if start <= today <= peak:
                return {"label": e["label"], "detail": e["detail"], "phase": "active",
                        "days_to_peak": (peak - today).days, "peak": e["peak"]}
            if today < start and (start - today).days <= lead:
                return {"label": e["label"], "detail": e["detail"], "phase": "upcoming",
                        "days_to_start": (start - today).days, "peak": e["peak"]}
        return None

    def _top_model(self) -> list[dict]:
        """Evaluate the creator's market-top checklist (fully automated sources)."""
        out = []
        for key, m in self.cfg.get("top_model", {}).items():
            src = m["source"]
            value, as_of, triggered = None, None, None
            try:
                if src == "fred_ratio":
                    value, as_of = self._fetch_fred_ratio(m)
                    triggered = value is not None and value >= m.get("trigger_above", 1e9)
                elif src == "fred_yoy_diff":
                    value, as_of = self._fetch_fred_yoy_diff(m)
                    triggered = value is not None and value <= m.get("trigger_below", -1e9)
                elif src == "ssga_top10":
                    value, as_of = self._fetch_ssga_top10(m)
                    triggered = value is not None and value >= m.get("trigger_above", 1e9)
            except Exception as e:  # noqa: BLE001
                logger.warning("Top-model %s failed: %s", key, e)
            out.append({
                "key": key,
                "label": m["label"],
                "unit": m.get("unit"),
                "value": value,
                "as_of": as_of,
                "next_expected": self._next_report_window() if m.get("quarterly") else None,
                "triggered": triggered,
                "live": True,
                "note": m.get("note", ""),
            })
        return out

    # ------------------------------------------------------------------- public
    def snapshot(self) -> dict:
        """Fetch every metric and return the full scored monitor state."""
        metrics = {}
        for key, m in self.cfg["metrics"].items():
            fetcher = getattr(self, self._FETCHERS[m["source"]])
            try:
                series = fetcher(m)
            except Exception as e:  # noqa: BLE001
                logger.error("Fetch failed for %s: %s", key, e)
                series = []
            metrics[key] = self._evaluate(key, m, series)

        # Enrich the quarterly capex metric with its latest-report timing.
        cap = metrics.get("cloud_capex_accel")
        if cap and cap.get("series"):
            last = cap["series"][-1]
            cap["as_of_quarter"] = self._quarter_label(last["date"])
            cap["as_of_filed"] = last.get("filed") or None
            cap["next_expected"] = self._next_report_window()

        confluence = self._evaluate_confluence(metrics)

        # The HEADLINE overall status follows the creator's confluence framework
        # (his alerts are combos, not any single card). "top" collapses to "red"
        # for the traffic-light. Individual cards keep their own status for detail.
        overall = {"top": "red", "red": "red", "yellow": "yellow", "normal": "normal"}[
            confluence["level"]
        ]

        # Per-card aggregate (worst non-context card) kept for reference/UI.
        statuses = [v["status"] for v in metrics.values() if not v.get("context_only")]
        if "red" in statuses:
            card_alert = "red"
        elif "yellow" in statuses:
            card_alert = "yellow"
        elif statuses and all(s == "unavailable" for s in statuses):
            card_alert = "unavailable"
        else:
            card_alert = "normal"

        # ---- Sell-now triggers: attach labels to metrics + compute snapshot flag.
        cc = self.cfg.get("confluence", {})
        rv = lambda k: metrics.get(k, {}).get("current")  # noqa: E731
        reserves, sofr, srf = rv("reserves"), rv("sofr_iorb_spread"), rv("srf_usage")
        sell_active_map = {
            "reserves": reserves is not None and reserves < cc.get("reserves_liquidate", 2500000),
            "cloud_capex_accel": confluence["capex_rolling_over"],
            "sofr_iorb_spread": (
                sofr is not None and reserves is not None
                and sofr > cc.get("sofr_red_bps", 3.0)
                and reserves < cc.get("reserves_red_break", 2800000)
            ),
            "srf_usage": self._srf_is_stress(srf, reserves, sofr, cc),
        }
        sell_reasons = []
        for key, m in metrics.items():
            ss = self.cfg["metrics"].get(key, {}).get("sell_signal")
            if not ss:
                continue
            active = bool(sell_active_map.get(key))
            m["sell_signal"] = ss
            m["sell_active"] = active
            if active:
                sell_reasons.append({
                    "key": key, "label": m["label"],
                    "priority": ss.get("priority"), "action": ss.get("action"),
                })
        # High-priority active OR confluence structural top = sell-now state.
        sell_now = any(r["priority"] == "high" for r in sell_reasons) or confluence["level"] in ("red", "top")

        top_model = self._top_model()
        top_triggered = sum(1 for t in top_model if t["triggered"])

        return {
            "generated_at": datetime.utcnow().isoformat() + "Z",
            "overall_status": overall,
            "card_alert_status": card_alert,
            "sell_now": sell_now,
            "sell_reasons": sell_reasons,
            "liquidity_calendar": self._liquidity_calendar(),
            "metrics": metrics,
            "confluence": confluence,
            "top_model": top_model,
            "top_model_triggered": top_triggered,
            "top_model_total": len(top_model),
        }


if __name__ == "__main__":  # manual smoke test
    snap = LiquidityMonitor().snapshot()
    print(f"OVERALL: {snap['overall_status']}\n")
    for k, v in snap["metrics"].items():
        cur = v.get("current")
        pct = v.get("percentile")
        print(
            f"{v['status']:12s} {v['label']:42s} "
            f"cur={cur} pct={pct} n={len(v['series'])}"
        )
